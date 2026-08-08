"""Gera snapshots Excel do Radar IA sem depender de ferramentas do Codex."""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def _number(value):
    return None if value in (None, "") else float(value)


def _style_header(sheet, range_reference):
    fill = PatternFill("solid", fgColor="102A43")
    for row in sheet[range_reference]:
        for cell in row:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_grid(sheet):
    border = Border(bottom=Side(style="thin", color="E2E8F0"))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border


def build_workbook(history_path: Path, output_dir: Path, run_id: str) -> Path:
    with history_path.open(newline="", encoding="utf-8") as source:
        rows = list(csv.DictReader(source))
    rows.sort(key=lambda row: (row["reference_month"], row["model_name"]))
    latest_month = rows[-1]["reference_month"]
    latest = [row for row in rows if row["reference_month"] == latest_month]
    latest.sort(key=lambda row: _number(row["intelligence_index"]) or float("-inf"), reverse=True)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Visão geral"
    current = workbook.create_sheet("Última coleta")
    history = workbook.create_sheet("Histórico")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False

    overview.merge_cells("A1:F1")
    overview["A1"] = "Radar IA — histórico de modelos"
    overview["A1"].fill = PatternFill("solid", fgColor="102A43")
    overview["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    overview["A1"].alignment = Alignment(horizontal="center")
    overview.row_dimensions[1].height = 30
    for index, (label, value) in enumerate([
        ("Mês mais recente", latest_month),
        ("Modelos na última coleta", len(latest)),
        ("Observações históricas", len(rows)),
        ("Versão do Intelligence Index", ", ".join(sorted({row["intelligence_index_version"] for row in latest if row["intelligence_index_version"]})) or "Não informada pela API"),
    ], start=3):
        overview.cell(index, 1, label).font = Font(bold=True, color="102A43")
        overview.cell(index, 1).fill = PatternFill("solid", fgColor="E8F1FA")
        overview.cell(index, 2, value)
    for index, (label, value) in enumerate([
        ("Fonte e uso", "Dados coletados da API gratuita da Artificial Analysis."),
        ("Metodologia", "O histórico preserva a versão do índice para evitar comparações silenciosas entre metodologias diferentes."),
        ("Atualização", "Execute python radar_ia.py uma vez por mês. O comando adiciona o mês novo e cria este Excel."),
        ("Atribuição", "https://artificialanalysis.ai/"),
    ], start=8):
        overview.cell(index, 1, label).font = Font(bold=True, color="102A43")
        overview.cell(index, 2, value).alignment = Alignment(wrap_text=True, vertical="top")
    overview.column_dimensions["A"].width = 31
    overview.column_dimensions["B"].width = 85

    current.append(["Posição", "Modelo", "Empresa", "Lançamento", "Intelligence", "Coding", "Agentic", "Speed (tok/s)", "Custo/tarefa (US$)", "Input US$/1M", "Output US$/1M", "Versão índice"])
    for position, row in enumerate(latest, start=1):
        current.append([position, row["model_name"], row["model_creator"], row["release_date"], _number(row["intelligence_index"]), _number(row["coding_index"]), _number(row["agentic_index"]), _number(row["median_output_tokens_per_second"]), _number(row["intelligence_index_cost_per_task_usd"]), _number(row["price_1m_input_tokens_usd"]), _number(row["price_1m_output_tokens_usd"]), row["intelligence_index_version"]])
    _style_header(current, "A1:L1")
    current.freeze_panes = "A2"
    for column, width in {"A": 10, "B": 38, "C": 20, "D": 14, "E": 14, "F": 12, "G": 12, "H": 15, "I": 18, "J": 15, "K": 16, "L": 16}.items():
        current.column_dimensions[column].width = width

    history.append(["Mês", "Coletado em UTC", "ID", "Slug", "Modelo", "Empresa", "Lançamento", "Tier", "Versão índice", "Intelligence", "Coding", "Agentic", "Speed (tok/s)", "TTFT (s)", "Custo/tarefa (US$)", "Input US$/1M", "Output US$/1M", "Fonte"])
    for row in rows:
        history.append([row["reference_month"], row["collected_at_utc"], row["model_id"], row["model_slug"], row["model_name"], row["model_creator"], row["release_date"], row["api_tier"], row["intelligence_index_version"], _number(row["intelligence_index"]), _number(row["coding_index"]), _number(row["agentic_index"]), _number(row["median_output_tokens_per_second"]), _number(row["median_time_to_first_token_seconds"]), _number(row["intelligence_index_cost_per_task_usd"]), _number(row["price_1m_input_tokens_usd"]), _number(row["price_1m_output_tokens_usd"]), row["source_url"]])
    _style_header(history, "A1:R1")
    history.freeze_panes = "A2"
    for column, width in {"A": 10, "B": 22, "C": 38, "D": 32, "E": 42, "F": 20, "G": 14, "H": 12, "I": 14, "J": 12, "K": 12, "L": 12, "M": 15, "N": 12, "O": 18, "P": 15, "Q": 16, "R": 52}.items():
        history.column_dimensions[column].width = width

    for sheet, date_column, currency_columns in ((current, "D", ("I", "J", "K")), (history, "G", ("O", "P", "Q"))):
        _apply_grid(sheet)
        for cell in sheet[date_column][1:]:
            cell.number_format = "yyyy-mm-dd"
        for column in currency_columns:
            for cell in sheet[column][1:]:
                cell.number_format = '$#,##0.000'

    output_dir.mkdir(exist_ok=True)
    output_path = output_dir / f"Radar_IA_{run_id}.xlsx"
    workbook.save(output_path)
    return output_path
